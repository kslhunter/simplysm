"""XLSX handler: extract cell data, images, and embedded objects.

Output format: per sheet, cell data is rendered as a markdown table whose
column headers are Excel column letters (A, B, C, ...) and whose first
column is the original Excel row number.  When an image is anchored to a
row, the current table chunk is flushed, the [IMG:N] placeholder is
emitted, and a new table (re-rendering the header) resumes from the next
row.  This preserves the spatial relationship between cell data and
images while keeping each chunk a valid markdown table that LLMs parse
natively.
"""

import zipfile
from _common import ensure_packages

PACKAGES = {"openpyxl": "openpyxl"}


def _escape_md(v):
    if v is None:
        return ""
    s = str(v).strip()
    return (
        s.replace("\\", "\\\\")
        .replace("|", "\\|")
        .replace("\r\n", "<br>")
        .replace("\n", "<br>")
        .replace("\r", "<br>")
    )


def _render_chunk(chunk_rows, max_col, get_col_letter):
    if not chunk_rows:
        return []
    headers = ["Row"] + [get_col_letter(c) for c in range(1, max_col + 1)]
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for row_num, cells in chunk_rows:
        padded = list(cells) + [""] * (max_col - len(cells))
        out.append(f"| {row_num} | " + " | ".join(padded[:max_col]) + " |")
    return out


def extract(file_path):
    ensure_packages(PACKAGES)
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet

    wb = load_workbook(file_path, data_only=True)
    text_parts = []
    images = []
    embedded = []
    img_idx = 0
    emb_idx = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"[Sheet: {sheet_name}]")
        text_parts.append("")

        if not isinstance(ws, Worksheet):
            text_parts.append(f"({type(ws).__name__} — 데이터 없음)")
            text_parts.append("")
            continue

        if ws.max_row is None or ws.max_row == 0:
            text_parts.append("(empty sheet)")
            text_parts.append("")
            continue

        # Merged cells annotation
        merged = list(ws.merged_cells.ranges)
        if merged:
            text_parts.append(f"[Merged: {', '.join(str(r) for r in merged)}]")
            text_parts.append("")

        ws_images = getattr(ws, "_images", [])
        row_img_markers = {}
        for img in ws_images:
            data_fn = getattr(img, "_data", None)
            blob = data_fn() if callable(data_fn) else b""
            if not blob:
                continue
            img_idx += 1
            anchor = getattr(img, "anchor", None)
            anchor_row = None
            anchor_col = None
            if anchor:
                _from = getattr(anchor, "_from", None)
                if _from:
                    anchor_row = getattr(_from, "row", None)
                    anchor_col = getattr(_from, "col", None)
                    if anchor_row is not None:
                        anchor_row += 1
                    if anchor_col is not None:
                        anchor_col += 1
            if anchor_row is None:
                anchor_row = ws.max_row or 1
            cell_ref = ""
            if anchor_col is not None:
                cell_ref = f" anchor:{get_column_letter(anchor_col)}{anchor_row}"
            else:
                cell_ref = f" anchor:row {anchor_row}"
            images.append({
                "data": blob,
                "ext": "png",
                "context": f"sheet '{sheet_name}'{cell_ref}",
            })
            row_img_markers.setdefault(anchor_row, []).append(img_idx)

        max_col = ws.max_column or 1
        chunk = []

        for row in ws.iter_rows(values_only=False):
            row_num = row[0].row
            cells = [_escape_md(c.value) for c in row]
            chunk.append((row_num, cells))

            if row_num in row_img_markers:
                text_parts.extend(_render_chunk(chunk, max_col, get_column_letter))
                text_parts.append("")
                for idx in row_img_markers[row_num]:
                    text_parts.append(f"[IMG:{idx}]")
                text_parts.append("")
                chunk = []

        if chunk:
            text_parts.extend(_render_chunk(chunk, max_col, get_column_letter))
            text_parts.append("")

    try:
        with zipfile.ZipFile(file_path, "r") as zf:
            for name in zf.namelist():
                if "embeddings/" in name.lower():
                    filename = name.split("/")[-1]
                    data = zf.read(name)
                    emb_idx += 1
                    embedded.append({"filename": filename, "data": data})
                    text_parts.append(f"[EMB:{emb_idx}]")
    except Exception:
        pass

    return {
        "text": "\n".join(text_parts),
        "images": images,
        "embedded": embedded,
        "metadata": {},
    }
