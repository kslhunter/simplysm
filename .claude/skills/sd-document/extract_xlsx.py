#!/usr/bin/env python3
"""Extract data and images from XLSX files with cell positions."""

import sys
from _common import (
    setup_encoding, make_output_paths, print_header, save_image,
    print_image_summary, run_cli,
)

setup_encoding()


def extract(file_path):
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    fp, out_dir = make_output_paths(file_path)
    img_idx = 0

    print_header(fp)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"## Sheet: {sheet_name}\n")

        # Data extraction
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            print("(empty sheet)\n")
            continue

        for row in rows:
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                else:
                    cells.append(str(val).strip())
            print(f"[{row[0].column_letter}{row[0].row}] " + " | ".join(cells))

        # Image extraction
        images = getattr(ws, '_images', [])
        if images:
            for img in images:
                img_idx += 1
                anchor = ""
                anchor_from = getattr(getattr(img, 'anchor', None), '_from', None)
                if anchor_from is not None:
                    anchor = f" (near {anchor_from.col},{anchor_from.row})"
                data_fn = getattr(img, '_data', None)
                blob = data_fn() if callable(data_fn) else b""
                img_path = save_image(out_dir, img_idx, blob, "png")
                print(f"[IMG]{anchor} {img_path}")

        print()

    print_image_summary(img_idx, out_dir)


if __name__ == "__main__":
    run_cli(extract, "extract_xlsx.py", {"openpyxl": "openpyxl"})
