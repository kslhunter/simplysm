#!/usr/bin/env python3
"""Extract data and images from XLSX files with cell positions."""

import sys
import io
import os
import subprocess
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")


def ensure_packages():
    packages = {"openpyxl": "openpyxl"}
    for pip_name, import_name in packages.items():
        try:
            __import__(import_name)
        except ImportError:
            print(f"Installing package: {pip_name}...", file=sys.stderr)
            subprocess.check_call([sys.executable, "-m", "pip", "install", pip_name],
                                  stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


def extract(file_path):
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XlImage

    wb = load_workbook(file_path, data_only=True)
    stem = Path(file_path).stem
    out_dir = Path(file_path).parent / f"{stem}_files"
    img_idx = 0

    print(f"# {Path(file_path).name}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"## Sheet: {sheet_name}\n")

        # Data extraction
        rows = list(ws.iter_rows(values_only=False))
        if not rows:
            print("(empty sheet)\n")
            continue

        for row in rows:
            cells = []
            for cell in row:
                val = cell.value
                if val is None:
                    cells.append("")
                else:
                    cells.append(str(val).strip())
            print(f"[{row[0].coordinate.split('1')[0]}{row[0].row}] " + " | ".join(cells))

        # Image extraction
        if ws._images:
            for img in ws._images:
                img_idx += 1
                anchor = ""
                if hasattr(img.anchor, '_from'):
                    anchor = f" (near {img.anchor._from.col},{img.anchor._from.row})"
                ext = "png"
                out_dir.mkdir(parents=True, exist_ok=True)
                img_path = out_dir / f"img_{img_idx:03d}.{ext}"
                with open(img_path, "wb") as f:
                    f.write(img._data())
                print(f"[IMG]{anchor} {img_path}")

        print()

    if img_idx > 0:
        print(f"---\n{img_idx} image(s) saved: {out_dir}")
    else:
        print("---\nNo images")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python extract_xlsx.py <file.xlsx>", file=sys.stderr)
        sys.exit(1)
    ensure_packages()
    extract(sys.argv[1])
